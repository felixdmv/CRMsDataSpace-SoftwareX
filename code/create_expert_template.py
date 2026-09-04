#!/usr/bin/env python3
"""
Generate updated expert templates for UNFC annotation.

Current business rule:
- One UNFC classification per mine / dump / tailings facility / deposit as a whole.
- No separate UNFC code by commodity.
- Experts justify the final E/F/G classification using as many variables as needed.
"""
from __future__ import annotations

import csv
import os
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
TEMPLATES_DIR = ROOT / "templates"
ES_DIR = TEMPLATES_DIR / "es"
EN_DIR = TEMPLATES_DIR / "en"


SOURCE_TYPE_OPTIONS_ES = (
    "escombrera|balsa de lodos|instalacion de relaves|deposito minero|mina|"
    "acopio|residuo metalurgico|escorial|otro"
)
SOURCE_TYPE_OPTIONS_EN = (
    "waste dump|sludge pond|tailings storage facility|mineral deposit|mine|stockpile|"
    "ore stockpile|metallurgical residue|slag heap|other"
)

STUDY_LEVEL_OPTIONS_ES = "sin_estudio|scoping|PEA|PFS|FS|operacion|solo_informe_tecnico|otro"
STUDY_LEVEL_OPTIONS_EN = "none|scoping|PEA|PFS|FS|operating|technical_report_only|other"
REVIEW_STATUS_ES = "pendiente|en_progreso|completado"
REVIEW_STATUS_EN = "pending|in_progress|completed"
QUESTION_TYPE_OPTIONS_ES = "tecnica|economica|ambiental|regulatoria|social|UNFC|clasificacion"
QUESTION_TYPE_OPTIONS_EN = "technical|economic|environmental|regulatory|social|UNFC|classification"
BASE_CANTIDAD_GUIDE_ES = "1=metal contenido o sustancia util contenida|2=producto vendible o recuperable|3=tonelaje bruto del material|4=otra base definida por el experto"
BASE_CANTIDAD_GUIDE_EN = "1=contained metal or contained useful substance|2=saleable or recoverable product|3=gross material tonnage|4=other basis defined by expert"

SHEET_THEME = {
    "LEEME": {"tab": "1D4ED8", "header": "1D4ED8", "soft": "DBEAFE"},
    "README": {"tab": "1D4ED8", "header": "1D4ED8", "soft": "DBEAFE"},
    "registro_informes": {"tab": "0F766E", "header": "0F766E", "soft": "CCFBF1"},
    "report_register": {"tab": "0F766E", "header": "0F766E", "soft": "CCFBF1"},
    "evaluacion_unfc": {"tab": "7C3AED", "header": "7C3AED", "soft": "EDE9FE"},
    "unfc_evaluation": {"tab": "7C3AED", "header": "7C3AED", "soft": "EDE9FE"},
    "evidencia_ejes": {"tab": "B45309", "header": "B45309", "soft": "FEF3C7"},
    "axis_evidence": {"tab": "B45309", "header": "B45309", "soft": "FEF3C7"},
    "pares_qa": {"tab": "BE185D", "header": "BE185D", "soft": "FCE7F3"},
    "qa_pairs": {"tab": "BE185D", "header": "BE185D", "soft": "FCE7F3"},
}

ROW_LABEL_EXAMPLE = {"EJEMPLO", "EXAMPLE"}
ROW_LABEL_EXPERTS = {"EXPERTOS", "EXPERTS"}


VARIABLE_GUIDE_ES = [
    ["eje", "variable", "descripcion", "escala_numerica", "notas"],
    ["E", "viabilidad_economica", "Viabilidad económica global del proyecto de recuperación.", "0 desconocido, 1 baja, 2 media, 3 alta", "Usar economía global del conjunto de materiales, no de un elemento aislado."],
    ["E", "demanda_mercado", "Demanda de mercado de los materiales o productos del conjunto.", "0 desconocida, 1 baja, 2 media, 3 alta", "Considerar mercado de salida, interés industrial y potencial de venta."],
    ["E", "valor_estimado", "Calidad de la estimación de valor o ingresos.", "0 ninguna, 1 aproximada, 2 preliminar, 3 detallada", "Mayor puntuación si el informe cuantifica valor o retorno."],
    ["E", "coste_estimado", "Calidad de la estimación de costes.", "0 ninguna, 1 aproximada, 2 preliminar, 3 detallada", "Considerar capex, opex, tratamiento, transporte, permisos y cierre."],
    ["E", "aceptacion_social", "Aceptación social o ausencia de conflicto relevante.", "0 desconocida, 1 baja, 2 media, 3 alta", "Usar consultas, oposición local, aceptación pública o apoyo social."],
    ["E", "area_protegida", "Restricción asociada a área protegida o figura equivalente.", "0 desconocida, 1 restricción alta, 2 restricción media, 3 restricción baja o nula", "Puntuar lo favorable que es la condición para la viabilidad."],
    ["E", "restriccion_ambiental", "Restricciones ambientales o legales.", "0 desconocida, 1 alta, 2 media, 3 baja", "Las puntuaciones bajas implican más barreras."],
    ["E", "impacto_ambiental", "Impacto ambiental esperado del proyecto.", "0 desconocido, 1 alto, 2 medio, 3 bajo", "Puntuar favorabilidad, no longitud del informe."],
    ["E", "estado_regulatorio", "Madurez de permisos y situación regulatoria.", "0 desconocido, 1 no iniciado, 2 en progreso, 3 autorizado o muy avanzado", "Usar permisos, concesiones, autorizaciones y condicionantes."],
    ["E", "permiso_requerido", "Carga de permisos adicionales pendiente.", "0 desconocida, 1 alta, 2 media, 3 baja o cubierta", "Mayor puntuación implica menor barrera regulatoria."],
    ["F", "madurez_proyecto", "Madurez global del proyecto.", "1 F1, 2 F2, 3 F3, 4 F4, 9 desconocido", "La categoría final F se registra en la hoja de evaluación."],
    ["F", "proyecto_definido", "Grado de definición del proyecto.", "0 ninguno, 1 conceptual, 2 parcial, 3 bien definido", "Usar definición de proceso, alcance, producto y plan de implantación."],
    ["F", "nivel_estudio", "Nivel de estudio más alto evidenciado.", "0 ninguno, 1 scoping, 2 PEA, 3 PFS/FS/operación", "Usar el nivel más alto claramente identificado en el informe."],
    ["F", "estado_desarrollo", "Estado de desarrollo del proyecto.", "0 desconocido, 1 temprano, 2 evaluación, 3 construcción u operación", "Mirar el estado del proyecto, no solo del activo minero."],
    ["G", "sondeos", "Soporte de sondeos o trabajos subsuperficiales.", "0 ninguno, 1 bajo, 2 medio, 3 alto", "Aplicado al conocimiento del conjunto del depósito o escombrera."],
    ["G", "muestreo", "Calidad y cobertura del muestreo.", "0 ninguno, 1 baja, 2 media, 3 alta", "Usar cantidad, representatividad y distribución del muestreo."],
    ["G", "composicion_quimica", "Confianza en la caracterización química.", "0 ninguna, 1 baja, 2 media, 3 alta", "Aplicada al conjunto de materiales y su posible valorización."],
    ["G", "propiedades_fisicas", "Confianza en la caracterización física/mineralógica.", "0 ninguna, 1 baja, 2 media, 3 alta", "Ejemplos: densidad, humedad, granulometría, mineralogía."],
    ["G", "distribucion_espacial", "Confianza en la geometría y distribución espacial.", "0 ninguna, 1 baja, 2 media, 3 alta", "Usar topografía, secciones, modelado, continuidad o cartografía."],
    ["G", "observaciones", "Soporte de observaciones de campo e interpretación.", "0 ninguno, 1 bajo, 2 medio, 3 alto", "Visitas, logs, fotos, observaciones directas e interpretación."],
    ["G", "criticidad", "Criticidad estratégica del conjunto de materiales según regulación, mercado o contexto europeo.", "0 desconocida, 1 baja, 2 media, 3 alta", "Nueva variable: permite reflejar importancia estratégica de los materiales."],
    ["E/F/G", "otra_variable", "Variable adicional definida por el experto si considera que la lista no es suficiente.", "usar criterio experto", "Permite añadir todas las variables necesarias para justificar la clasificación final."],
]

VARIABLE_GUIDE_EN = [
    ["axis", "variable", "description", "numeric_scale", "notes"],
    ["E", "economic_viability", "Overall economic viability of the recovery project.", "0 unknown, 1 low, 2 medium, 3 high", "Use the project as a whole, not one isolated commodity."],
    ["E", "market_demand", "Market demand for the overall material/product set.", "0 unknown, 1 low, 2 medium, 3 high", "Consider market outlet, industrial demand, and sales potential."],
    ["E", "estimated_value", "Quality of value or revenue estimate.", "0 none, 1 rough, 2 preliminary, 3 detailed", "Higher score if the report quantifies value or return."],
    ["E", "estimated_cost", "Quality of cost estimate.", "0 none, 1 rough, 2 preliminary, 3 detailed", "Consider capex, opex, treatment, transport, permits, and closure."],
    ["E", "social_acceptance", "Social acceptance or absence of relevant conflict.", "0 unknown, 1 low, 2 medium, 3 high", "Use consultations, local opposition, public acceptance, or support."],
    ["E", "protected_area", "Constraint linked to protected area or similar figure.", "0 unknown, 1 high restriction, 2 medium restriction, 3 low or no restriction", "Score how favourable the condition is for viability."],
    ["E", "environmental_restriction", "Environmental or legal restriction.", "0 unknown, 1 high, 2 medium, 3 low", "Lower scores mean more barriers."],
    ["E", "environmental_impact", "Expected environmental impact of the project.", "0 unknown, 1 high, 2 medium, 3 low", "Score favourability, not document length."],
    ["E", "regulatory_status", "Maturity of permits and regulatory status.", "0 unknown, 1 not started, 2 in progress, 3 authorized or advanced", "Use permits, concessions, approvals, and conditions."],
    ["E", "permit_required", "Remaining permitting burden.", "0 unknown, 1 high, 2 medium, 3 low or covered", "Higher score means lower permitting barrier."],
    ["F", "project_maturity", "Overall project maturity.", "1 F1, 2 F2, 3 F3, 4 F4, 9 unknown", "The final F category is recorded in the evaluation sheet."],
    ["F", "project_defined", "How clearly the project is defined.", "0 none, 1 conceptual, 2 partial, 3 well defined", "Use process route, scope, product, and implementation plan."],
    ["F", "study_level", "Highest study level evidenced.", "0 none, 1 scoping, 2 PEA, 3 PFS/FS/operating", "Use the highest clearly identified level."],
    ["F", "development_status", "Project development status.", "0 unknown, 1 early, 2 evaluation, 3 construction or operating", "Look at the project, not only the mining asset."],
    ["G", "boreholes", "Support from boreholes or subsurface works.", "0 none, 1 low, 2 medium, 3 high", "Applied to knowledge of the source as a whole."],
    ["G", "sampling", "Quality and coverage of sampling.", "0 none, 1 low, 2 medium, 3 high", "Use amount, representativeness, and coverage."],
    ["G", "chemical_composition", "Confidence in chemical characterization.", "0 none, 1 low, 2 medium, 3 high", "Applied to the material set and its valorisation potential."],
    ["G", "physical_properties", "Confidence in physical/mineralogical characterization.", "0 none, 1 low, 2 medium, 3 high", "Examples: density, moisture, grain size, mineralogy."],
    ["G", "spatial_distribution", "Confidence in geometry and spatial distribution.", "0 none, 1 low, 2 medium, 3 high", "Use topography, sections, modelling, continuity, or mapping."],
    ["G", "observations", "Support from field observations and interpretation.", "0 none, 1 low, 2 medium, 3 high", "Site visits, logs, photos, direct observations, interpretation."],
    ["G", "criticality", "Strategic criticality of the material set according to European regulation, market, or policy context.", "0 unknown, 1 low, 2 medium, 3 high", "New variable to reflect strategic importance of the materials."],
    ["E/F/G", "other_variable", "Additional variable defined by the expert if the list is not enough.", "expert-defined", "Allows experts to add all variables needed to justify the final classification."],
]


REPORT_COLUMNS_ES = [
    "tipo_fila",
    "nombre_pdf",
    "titulo_informe",
    "empresa",
    "pais",
    "region",
    "tipo_fuente",
    "nombre_fuente",
    "material_fuente",
    "elementos_observados",
    "nivel_estudio_reportado",
    "fecha_efectiva_o_ano",
    "nombre_experto",
    "estado_revision",
    "(notas)",
]

REPORT_COLUMNS_EN = [
    "row_type",
    "pdf_name",
    "report_title",
    "company",
    "country",
    "region",
    "source_type",
    "source_name",
    "source_material",
    "commodities_observed",
    "study_level_reported",
    "effective_date_or_year",
    "expert_name",
    "review_status",
    "(notes)",
]

EVAL_COLUMNS_ES = [
    "tipo_fila",
    "nombre_pdf",
    "nombre_fuente",
    "codigo_unfc_final",
    "categoria_e_final",
    "categoria_f_final",
    "categoria_g_final",
    "codigo_base_cantidad",
    "cantidad_medida",
    "cantidad_indicada",
    "cantidad_inferida",
    "unidad",
    "paginas_clave",
    "justificacion_experto",
]

EVAL_COLUMNS_EN = [
    "row_type",
    "pdf_name",
    "source_name",
    "unfc_code_final",
    "e_category_final",
    "f_category_final",
    "g_category_final",
    "quantity_basis_code",
    "measured_quantity",
    "indicated_quantity",
    "inferred_quantity",
    "unit",
    "key_pages",
    "expert_rationale",
]

EVIDENCE_COLUMNS_ES = [
    "tipo_fila",
    "nombre_pdf",
    "nombre_fuente",
    "eje",
    "variable",
    "valor_numerico",
    "paginas_evidencia",
    "(extracto_o_nota)",
    "(comentario_experto)",
]

EVIDENCE_COLUMNS_EN = [
    "row_type",
    "pdf_name",
    "source_name",
    "axis",
    "variable",
    "numeric_value",
    "evidence_pages",
    "(excerpt_or_note)",
    "(expert_comment)",
]

QA_COLUMNS_ES = [
    "tipo_fila",
    "nombre_pdf",
    "nombre_fuente",
    "tipo_pregunta",
    "pregunta",
    "respuesta_gold_corta",
    "respuesta_gold_larga",
    "terminos_obligatorios",
    "(alternativas_aceptables)",
    "paginas_fuente",
    "(notas_revisor)",
]

QA_COLUMNS_EN = [
    "row_type",
    "pdf_name",
    "source_name",
    "question_type",
    "question",
    "gold_answer_short",
    "gold_answer_long",
    "must_mention_terms",
    "(acceptable_alternatives)",
    "source_pages",
    "(reviewer_notes)",
]


REPORT_EXAMPLES = [
    ["EJEMPLO", "1._Cobre_Las_Cruces_Technical_Report_Feb_2024_FINAL.pdf", "Cobre Las Cruces: proyecto de sulfuros primarios polimetálicos", "First Quantum Minerals Ltd", "España", "Andalucía", "mina", "Cobre Las Cruces PMS / PMR", "sulfuros primarios polimetálicos", "cobre|zinc|plomo|plata", "solo_informe_tecnico", "2024", "Ejemplo de plantilla", "completado", "Ejemplo inventado completo para mostrar el formato."],
    ["EJEMPLO", "DMET_-_240523_Aguablanca_Technical_Report_PFS_Final.pdf", "Informe técnico y estudio de prefactibilidad preliminar del depósito de níquel-cobre de Aguablanca", "Denarius Metals Corp.", "España", "Extremadura", "mina", "Aguablanca", "depósito mineral de níquel-cobre", "níquel|cobre", "PFS", "2024", "Ejemplo de plantilla", "completado", "Ejemplo inventado completo para mostrar el formato."],
    ["EJEMPLO", "070912ElValleTechnicalReport.pdf", "Informe técnico de los depósitos auríferos El Valle, Carlés, La Brueva y Godán", "Buffalo Gold Ltd.", "España", "Asturias", "mina", "El Valle / Carlés / La Brueva / Godán", "depósitos de oro", "oro", "solo_informe_tecnico", "2007", "Ejemplo de plantilla", "completado", "Ejemplo inventado completo para mostrar el formato."],
    ["EXPERTOS", "NOMBRE_PDF_AQUI", "TITULO_INFORME_AQUI", "EMPRESA_AQUI", "PAIS_AQUI", "REGION_AQUI", "TIPO_FUENTE_AQUI", "NOMBRE_FUENTE_AQUI", "MATERIAL_FUENTE_AQUI", "ELEMENTOS_OBSERVADOS_AQUI", "NIVEL_ESTUDIO_AQUI", "ANO_O_FECHA_AQUI", "NOMBRE_EXPERTO_AQUI", "pendiente", "ESCRIBIR_NOTAS_AQUI"],
]

REPORT_EXAMPLES_EN = [
    ["EXAMPLE", "1._Cobre_Las_Cruces_Technical_Report_Feb_2024_FINAL.pdf", "Cobre Las Cruces: Polymetallic Primary Sulphide Project", "First Quantum Minerals Ltd", "Spain", "Andalusia", "mine", "Cobre Las Cruces PMS / PMR", "polymetallic primary sulphides", "copper|zinc|lead|silver", "technical_report_only", "2024", "Template example", "completed", "Complete invented example to show the format."],
    ["EXAMPLE", "DMET_-_240523_Aguablanca_Technical_Report_PFS_Final.pdf", "Technical Report and Preliminary Feasibility Study for the Aguablanca Nickel-Copper Mineral Deposit", "Denarius Metals Corp.", "Spain", "Extremadura", "mine", "Aguablanca", "nickel-copper mineral deposit", "nickel|copper", "PFS", "2024", "Template example", "completed", "Complete invented example to show the format."],
    ["EXAMPLE", "070912ElValleTechnicalReport.pdf", "Technical Report for the El Valle, Carlés, La Brueva, and Godán Gold Deposits", "Buffalo Gold Ltd.", "Spain", "Asturias", "mine", "El Valle / Carlés / La Brueva / Godán", "gold deposits", "gold", "technical_report_only", "2007", "Template example", "completed", "Complete invented example to show the format."],
    ["EXPERTS", "PDF_NAME_HERE", "REPORT_TITLE_HERE", "COMPANY_HERE", "COUNTRY_HERE", "REGION_HERE", "SOURCE_TYPE_HERE", "SOURCE_NAME_HERE", "SOURCE_MATERIAL_HERE", "COMMODITIES_OBSERVED_HERE", "STUDY_LEVEL_HERE", "YEAR_OR_DATE_HERE", "EXPERT_NAME_HERE", "pending", "WRITE_NOTES_HERE"],
]


EVAL_EXAMPLES_ES = [
    ["EJEMPLO", "DMET_-_240523_Aguablanca_Technical_Report_PFS_Final.pdf", "Aguablanca", "E1;F2;G2", "E1", "F2", "G2", "1", "15700", "25500", "16000", "t", "1-3,45-60", "Clasificación inventada completa del conjunto del proyecto."],
    ["EJEMPLO", "1._Cobre_Las_Cruces_Technical_Report_Feb_2024_FINAL.pdf", "Cobre Las Cruces PMS / PMR", "E1;F1;G2", "E1", "F1", "G2", "2", "35000", "42000", "14000", "t", "15-17,80-110", "Clasificación inventada completa del conjunto de materiales."],
    ["EJEMPLO", "070912ElValleTechnicalReport.pdf", "El Valle / Carlés / La Brueva / Godán", "E1;F2;G1", "E1", "F2", "G1", "1", "520000", "340000", "110000", "oz", "1-2,56-90", "Clasificación inventada completa del conjunto del activo."],
    ["EXPERTOS", "NOMBRE_PDF_AQUI", "NOMBRE_FUENTE_AQUI", "E1;F2;G3", "E1", "F2", "G3", "1", "CANTIDAD_MEDIDA_AQUI", "CANTIDAD_INDICADA_AQUI", "CANTIDAD_INFERIDA_AQUI", "UNIDAD_AQUI", "PAGINAS_AQUI", "JUSTIFICACION_AQUI"],
]

EVAL_EXAMPLES_EN = [
    ["EXAMPLE", "DMET_-_240523_Aguablanca_Technical_Report_PFS_Final.pdf", "Aguablanca", "E1;F2;G2", "E1", "F2", "G2", "1", "15700", "25500", "16000", "t", "1-3,45-60", "Complete invented classification for the project as a whole."],
    ["EXAMPLE", "1._Cobre_Las_Cruces_Technical_Report_Feb_2024_FINAL.pdf", "Cobre Las Cruces PMS / PMR", "E1;F1;G2", "E1", "F1", "G2", "2", "35000", "42000", "14000", "t", "15-17,80-110", "Complete invented classification for the material set as a whole."],
    ["EXAMPLE", "070912ElValleTechnicalReport.pdf", "El Valle / Carlés / La Brueva / Godán", "E1;F2;G1", "E1", "F2", "G1", "1", "520000", "340000", "110000", "oz", "1-2,56-90", "Complete invented classification for the asset as a whole."],
    ["EXPERTS", "PDF_NAME_HERE", "SOURCE_NAME_HERE", "E1;F2;G3", "E1", "F2", "G3", "1", "MEASURED_QUANTITY_HERE", "INDICATED_QUANTITY_HERE", "INFERRED_QUANTITY_HERE", "UNIT_HERE", "PAGES_HERE", "RATIONALE_HERE"],
]


def build_evidence_examples(lang: str) -> list[list[str]]:
    eval_rows = EVAL_EXAMPLES_ES if lang == "es" else EVAL_EXAMPLES_EN
    examples = []
    for row in eval_rows[:-1]:
        row_type, pdf_name, source_name = row[0], row[1], row[2]
        e_code, f_code, g_code = row[4], row[5], row[6]
        e_num = "3" if e_code == "E1" else ("2" if e_code == "E2" else "1")
        f_num = {"F1": "1", "F2": "2", "F3": "3", "F4": "4"}.get(f_code, "9")
        g_num = {"G1": "3", "G2": "2", "G3": "1", "G4": "0"}.get(g_code, "0")
        if lang == "es":
            examples.extend([
                [row_type, pdf_name, source_name, "E", "viabilidad_economica", e_num, row[12], "Variable destacada por el experto para justificar E.", "Ejemplo inventado."],
                [row_type, pdf_name, source_name, "E", "estado_regulatorio", e_num, row[12], "Otra variable E relevante para justificar la clase.", "Ejemplo inventado."],
                [row_type, pdf_name, source_name, "F", "nivel_estudio", f_num, row[12], "Variable destacada por el experto para justificar F.", "Ejemplo inventado."],
                [row_type, pdf_name, source_name, "F", "proyecto_definido", "3", row[12], "Otra variable F relevante para justificar la clase.", "Ejemplo inventado."],
                [row_type, pdf_name, source_name, "G", "muestreo", g_num, row[12], "Variable destacada por el experto para justificar G.", "Ejemplo inventado."],
                [row_type, pdf_name, source_name, "G", "criticidad", "2", row[12], "Nueva variable G: criticidad estratégica del conjunto de materiales.", "Ejemplo inventado."],
            ])
        else:
            examples.extend([
                [row_type, pdf_name, source_name, "E", "economic_viability", e_num, row[12], "Variable highlighted by the expert to justify E.", "Invented example."],
                [row_type, pdf_name, source_name, "E", "regulatory_status", e_num, row[12], "Another E variable relevant to the class.", "Invented example."],
                [row_type, pdf_name, source_name, "F", "study_level", f_num, row[12], "Variable highlighted by the expert to justify F.", "Invented example."],
                [row_type, pdf_name, source_name, "F", "project_defined", "3", row[12], "Another F variable relevant to the class.", "Invented example."],
                [row_type, pdf_name, source_name, "G", "sampling", g_num, row[12], "Variable highlighted by the expert to justify G.", "Invented example."],
                [row_type, pdf_name, source_name, "G", "criticality", "2", row[12], "New G variable: strategic criticality of the material set.", "Invented example."],
            ])
    if lang == "es":
        examples.append(["EXPERTOS", "NOMBRE_PDF_AQUI", "NOMBRE_FUENTE_AQUI", "E_F_O_G", "VARIABLE_AQUI", "VALOR_NUMERICO_AQUI", "PAGINAS_AQUI", "EXTRACTO_O_NOTA_AQUI", "COMENTARIO_EXPERTO_AQUI"])
    else:
        examples.append(["EXPERTS", "PDF_NAME_HERE", "SOURCE_NAME_HERE", "E_F_OR_G", "VARIABLE_HERE", "NUMERIC_VALUE_HERE", "PAGES_HERE", "EXCERPT_OR_NOTE_HERE", "EXPERT_COMMENT_HERE"])
    return examples


QA_EXAMPLES_ES = [
    ["EJEMPLO", "1._Cobre_Las_Cruces_Technical_Report_Feb_2024_FINAL.pdf", "Cobre Las Cruces PMS / PMR", "regulatoria", "Which metals are permitted in the concession modification?", "Copper, zinc, lead, and silver.", "The report summary states that copper, zinc, lead, and silver are included.", "copper|zinc|lead|silver", "cu|zn|pb|ag", "17", "Ejemplo inventado."],
    ["EJEMPLO", "DMET_-_240523_Aguablanca_Technical_Report_PFS_Final.pdf", "Aguablanca", "UNFC", "What study level is shown on the title page?", "Preliminary Feasibility Study (PFS).", "The title page identifies a Preliminary Feasibility Study.", "preliminary feasibility study|pfs", "pre-feasibility study", "1", "Ejemplo inventado."],
    ["EJEMPLO", "070912ElValleTechnicalReport.pdf", "El Valle / Carlés / La Brueva / Godán", "tecnica", "What is the main commodity in the report title?", "Gold.", "The title identifies the deposits as gold deposits.", "gold", "au", "1", "Ejemplo inventado."],
    ["EXPERTOS", "NOMBRE_PDF_AQUI", "NOMBRE_FUENTE_AQUI", "TIPO_PREGUNTA_AQUI", "ESCRIBIR_PREGUNTA_EN_INGLES", "RESPUESTA_GOLD_CORTA_AQUI", "RESPUESTA_GOLD_LARGA_AQUI", "TERMINOS_OBLIGATORIOS_AQUI", "ALTERNATIVAS_ACEPTABLES_AQUI", "PAGINAS_FUENTE_AQUI", "NOTAS_AQUI"],
]

QA_EXAMPLES_EN = [
    ["EXAMPLE", "1._Cobre_Las_Cruces_Technical_Report_Feb_2024_FINAL.pdf", "Cobre Las Cruces PMS / PMR", "regulatory", "Which metals are permitted in the concession modification?", "Copper, zinc, lead, and silver.", "The report summary states that copper, zinc, lead, and silver are included.", "copper|zinc|lead|silver", "cu|zn|pb|ag", "17", "Invented example."],
    ["EXAMPLE", "DMET_-_240523_Aguablanca_Technical_Report_PFS_Final.pdf", "Aguablanca", "UNFC", "What study level is shown on the title page?", "Preliminary Feasibility Study (PFS).", "The title page identifies a Preliminary Feasibility Study.", "preliminary feasibility study|pfs", "pre-feasibility study", "1", "Invented example."],
    ["EXAMPLE", "070912ElValleTechnicalReport.pdf", "El Valle / Carlés / La Brueva / Godán", "technical", "What is the main commodity in the report title?", "Gold.", "The title identifies the deposits as gold deposits.", "gold", "au", "1", "Invented example."],
    ["EXPERTS", "PDF_NAME_HERE", "SOURCE_NAME_HERE", "QUESTION_TYPE_HERE", "WRITE_QUESTION_IN_ENGLISH", "SHORT_GOLD_ANSWER_HERE", "LONG_GOLD_ANSWER_HERE", "MUST_MENTION_TERMS_HERE", "ACCEPTABLE_ALTERNATIVES_HERE", "SOURCE_PAGES_HERE", "NOTES_HERE"],
]


def rows_to_csv(path: Path, rows: Iterable[Iterable[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        csv.writer(handle).writerows(rows)


def find_row(rows: list[list[str]], marker: str) -> int | None:
    for idx, row in enumerate(rows, start=1):
        if row and row[0] == marker:
            return idx
    return None


def max_widths(rows: list[list[str]]) -> list[int]:
    width_map: list[int] = []
    for row in rows:
        for idx, value in enumerate(row):
            text = "" if value is None else str(value)
            if idx >= len(width_map):
                width_map.append(len(text))
            else:
                width_map[idx] = max(width_map[idx], len(text))
    return width_map


def sheet_header_row(sheet_name: str, rows: list[list[str]]) -> int | None:
    header_by_sheet = {
        "LEEME": "hoja_o_seccion",
        "README": "sheet_or_section",
        "registro_informes": "tipo_fila",
        "report_register": "row_type",
        "evaluacion_unfc": "tipo_fila",
        "unfc_evaluation": "row_type",
        "evidencia_ejes": "tipo_fila",
        "axis_evidence": "row_type",
        "pares_qa": "tipo_fila",
        "qa_pairs": "row_type",
    }
    header_marker = header_by_sheet.get(sheet_name)
    if not header_marker:
        return None
    match = None
    for idx, row in enumerate(rows, start=1):
        if row and row[0] == header_marker:
            match = idx
    return match


def add_list_validation(ws, formula_text: str, target_range: str) -> None:
    validation = DataValidation(type="list", formula1=f'"{formula_text}"', allow_blank=True)
    validation.prompt = "Choose one of the suggested values when possible."
    validation.error = "Please pick one of the suggested values or leave the field empty if allowed."
    ws.add_data_validation(validation)
    validation.add(target_range)


def apply_validations(ws, sheet_name: str, header_row: int, max_row: int, max_col: int) -> None:
    headers = {ws.cell(header_row, col).value: col for col in range(1, max_col + 1)}
    row_start = header_row + 1
    row_end = max(max_row + 50, 250)

    def target(column_name: str) -> str | None:
        col = headers.get(column_name)
        if not col:
            return None
        return f"{get_column_letter(col)}{row_start}:{get_column_letter(col)}{row_end}"

    validations = []
    if sheet_name == "registro_informes":
        validations = [
            ("tipo_fila", "EJEMPLO,EXPERTOS"),
            ("tipo_fuente", SOURCE_TYPE_OPTIONS_ES.replace("|", ",")),
            ("nivel_estudio_reportado", STUDY_LEVEL_OPTIONS_ES.replace("|", ",")),
            ("estado_revision", REVIEW_STATUS_ES.replace("|", ",")),
        ]
    elif sheet_name == "report_register":
        validations = [
            ("row_type", "EXAMPLE,EXPERTS"),
            ("source_type", SOURCE_TYPE_OPTIONS_EN.replace("|", ",")),
            ("study_level_reported", STUDY_LEVEL_OPTIONS_EN.replace("|", ",")),
            ("review_status", REVIEW_STATUS_EN.replace("|", ",")),
        ]
    elif sheet_name == "evaluacion_unfc":
        validations = [
            ("tipo_fila", "EJEMPLO,EXPERTOS"),
            ("categoria_e_final", "E1,E2,E3,E4,unknown"),
            ("categoria_f_final", "F1,F2,F3,F4,unknown"),
            ("categoria_g_final", "G1,G2,G3,G4,unknown"),
            ("codigo_base_cantidad", "1,2,3,4"),
        ]
    elif sheet_name == "unfc_evaluation":
        validations = [
            ("row_type", "EXAMPLE,EXPERTS"),
            ("e_category_final", "E1,E2,E3,E4,unknown"),
            ("f_category_final", "F1,F2,F3,F4,unknown"),
            ("g_category_final", "G1,G2,G3,G4,unknown"),
            ("quantity_basis_code", "1,2,3,4"),
        ]
    elif sheet_name == "evidencia_ejes":
        validations = [
            ("tipo_fila", "EJEMPLO,EXPERTOS"),
            ("eje", "E,F,G"),
        ]
    elif sheet_name == "axis_evidence":
        validations = [
            ("row_type", "EXAMPLE,EXPERTS"),
            ("axis", "E,F,G"),
        ]
    elif sheet_name == "pares_qa":
        validations = [
            ("tipo_fila", "EJEMPLO,EXPERTOS"),
            ("tipo_pregunta", QUESTION_TYPE_OPTIONS_ES.replace("|", ",")),
        ]
    elif sheet_name == "qa_pairs":
        validations = [
            ("row_type", "EXAMPLE,EXPERTS"),
            ("question_type", QUESTION_TYPE_OPTIONS_EN.replace("|", ",")),
        ]

    for column_name, formula_text in validations:
        target_range = target(column_name)
        if target_range:
            add_list_validation(ws, formula_text, target_range)


def format_sheet(ws, sheet_name: str, rows: list[list[str]]) -> None:
    theme = SHEET_THEME.get(sheet_name, {"tab": "334155", "header": "334155", "soft": "E5E7EB"})
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = theme["tab"]
    ws.sheet_view.zoomScale = 90

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    base_align = Alignment(vertical="top", wrap_text=True)
    center_align = Alignment(vertical="top", horizontal="center", wrap_text=True)

    dark_fill = PatternFill("solid", fgColor=theme["header"])
    soft_fill = PatternFill("solid", fgColor=theme["soft"])
    example_fill = PatternFill("solid", fgColor="F8FAFC")
    experts_fill = PatternFill("solid", fgColor="FFF7ED")
    section_fill = PatternFill("solid", fgColor="F8FAFC")
    axis_fill = PatternFill("solid", fgColor="F8FAFC")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    normal_font = Font(color="111827")

    for row_idx, row in enumerate(rows, start=1):
        first_value = row[0] if row else ""
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = base_align
            cell.border = border
            cell.font = normal_font
            if value == "":
                cell.value = None
            if col_idx == 1 and first_value in ROW_LABEL_EXAMPLE | ROW_LABEL_EXPERTS:
                cell.alignment = center_align

        if not row or all(str(v).strip() == "" for v in row):
            continue

        if row_idx == 1:
            for cell in ws[row_idx]:
                cell.fill = dark_fill
                cell.font = white_font
                cell.alignment = center_align
            ws.row_dimensions[row_idx].height = 24
            continue

        if first_value in ("campo", "field", "eje", "axis"):
            for cell in ws[row_idx]:
                cell.fill = section_fill
                cell.font = bold_font
                cell.alignment = center_align
            ws.row_dimensions[row_idx].height = 24
            continue

        if first_value in ROW_LABEL_EXAMPLE:
            ws.cell(row=row_idx, column=1).fill = example_fill
            ws.cell(row=row_idx, column=1).font = Font(bold=True, color="475569")
            continue

        if first_value in ROW_LABEL_EXPERTS:
            ws.cell(row=row_idx, column=1).fill = experts_fill
            ws.cell(row=row_idx, column=1).font = Font(bold=True, color="9A3412")
            continue

        if row_idx == sheet_header_row(sheet_name, rows):
            for cell in ws[row_idx]:
                cell.fill = dark_fill
                cell.font = white_font
                cell.alignment = center_align
            ws.row_dimensions[row_idx].height = 26
            continue

        if row_idx < (sheet_header_row(sheet_name, rows) or 0):
            ws.row_dimensions[row_idx].height = 32

    widths = max_widths(rows)
    for idx, width in enumerate(widths, start=1):
        adjusted = min(max(width + 3, 14), 42)
        ws.column_dimensions[get_column_letter(idx)].width = adjusted

    header_row = sheet_header_row(sheet_name, rows)
    if header_row:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
        apply_validations(ws, sheet_name, header_row, ws.max_row, ws.max_column)


def build_xlsx(path: Path, sheets: list[tuple[str, list[list[str]]]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)

    for name, rows in sheets:
        ws = workbook.create_sheet(title=name)
        for row in rows:
            ws.append(row)
        format_sheet(ws, name, rows)

    workbook.save(path)


def readme_rows(lang: str) -> list[list[str]]:
    if lang == "es":
        return [
            ["hoja_o_seccion", "que_es", "que_hay_que_rellenar", "desde_que_fila"],
            ["registro_informes", "Una fila por PDF.", "Rellenar metadatos del informe y del activo. Las filas EJEMPLO ya están completas. Empezar en la fila EXPERTOS.", "fila 22"],
            ["evaluacion_unfc", "Una fila por mina, escombrera, depósito o activo.", "Rellenar una única clasificación UNFC para el conjunto del activo.", "fila 21"],
            ["evidencia_ejes", "Una fila por variable usada para justificar E, F o G.", "Añadir todas las variables que el experto considere necesarias para justificar la clasificación.", "fila 55"],
            ["pares_qa", "Preguntas y respuestas gold.", "Rellenar preguntas en inglés y respuestas gold en inglés.", "fila 18"],
            ["regla_clave", "No separar UNFC por elemento.", "La clasificación se hace para el conjunto del activo y materiales, no por elemento individual.", ""],
            ["objetivo", "Pedir lo mínimo útil para comparar expertos frente a modelos.", "Los ejemplos están inventados y solo muestran el formato deseado.", ""],
        ]
    return [
        ["sheet_or_section", "what_it_is", "what_to_fill", "start_from_row"],
        ["report_register", "One row per PDF.", "Fill report and asset metadata. EXAMPLE rows are already completed. Start on the EXPERTS row.", "row 22"],
        ["unfc_evaluation", "One row per mine, dump, deposit, or asset.", "Fill one single UNFC classification for the asset as a whole.", "row 21"],
        ["axis_evidence", "One row per variable used to justify E, F, or G.", "Add as many variables as the expert needs to justify the final classification.", "row 55"],
        ["qa_pairs", "Gold questions and answers.", "Write questions in English and gold answers in English.", "row 18"],
        ["key_rule", "Do not separate UNFC by commodity.", "Classification applies to the asset and material set as a whole.", ""],
        ["purpose", "Ask the minimum useful information to compare experts and models.", "Examples are invented and only show the intended format.", ""],
    ]


def report_rows(lang: str) -> list[list[str]]:
    if lang == "es":
        guide = [
            ["campo", "formato", "opciones_o_ejemplo", "notas"],
            ["tipo_fila", "texto", "EJEMPLO|EXPERTOS", "Marca filas de ejemplo y la primera fila a completar por expertos."],
            ["nombre_pdf", "texto", "070912ElValleTechnicalReport.pdf", "Clave principal para cruzar hojas."],
            ["titulo_informe", "texto", "Titulo completo del informe", "Copiar el titulo visible en portada o metadatos."],
            ["empresa", "texto", "Nombre de la empresa o titular", "Si no aplica, indicar desconocido."],
            ["pais", "texto", "España|Portugal|Finlandia|otro", "Usar el pais principal del activo o del informe."],
            ["region", "texto", "Andalucia|Asturias|Extremadura|otra", "Region administrativa o geografica relevante."],
            ["tipo_fuente", "texto", SOURCE_TYPE_OPTIONS_ES, "Usar una de estas opciones siempre que sea posible."],
            ["nombre_fuente", "texto", "Nombre de mina, escombrera o deposito", "Debe coincidir en todas las hojas."],
            ["material_fuente", "texto", "relaves de flotacion|sulfuros|estériles|mineral oxidado|otro", "Describir el tipo general de material del activo."],
            ["elementos_observados", "texto", "niquel|cobre|oro", "Separar con barra vertical |."],
            ["nivel_estudio_reportado", "texto", STUDY_LEVEL_OPTIONS_ES, "Copiar el nivel de estudio visible en el informe."],
            ["fecha_efectiva_o_ano", "texto", "2024|2007|2024-02-15", "Año o fecha efectiva si aparece."],
            ["nombre_experto", "texto", "Nombre y apellido", "Persona que rellena o revisa la fila."],
            ["estado_revision", "texto", REVIEW_STATUS_ES, "Seguimiento mínimo del estado de revisión."],
            ["(notas)", "texto opcional", "Observaciones libres", "Opcional."],
            [""],
        ]
        return guide + [REPORT_COLUMNS_ES] + REPORT_EXAMPLES
    guide = [
        ["field", "format", "allowed_values_or_example", "notes"],
        ["row_type", "text", "EXAMPLE|EXPERTS", "Marks sample rows and the first row to be filled by experts."],
        ["pdf_name", "text", "070912ElValleTechnicalReport.pdf", "Primary key across sheets."],
        ["report_title", "text", "Full report title", "Copy the title visible on the cover or metadata."],
        ["company", "text", "Company or title holder name", "If unknown, write unknown."],
        ["country", "text", "Spain|Portugal|Finland|other", "Use the main country of the asset or report."],
        ["region", "text", "Andalusia|Asturias|Extremadura|other", "Relevant administrative or geographic region."],
        ["source_type", "text", SOURCE_TYPE_OPTIONS_EN, "Use one of these options whenever possible."],
        ["source_name", "text", "Name of the mine, dump, or deposit", "Must match across sheets."],
        ["source_material", "text", "flotation tailings|sulphides|waste rock|oxide ore|other", "General material type of the asset."],
        ["commodities_observed", "text", "nickel|copper|gold", "Separate values with |."],
        ["study_level_reported", "text", STUDY_LEVEL_OPTIONS_EN, "Copy the study level visible in the report."],
        ["effective_date_or_year", "text", "2024|2007|2024-02-15", "Year or effective date if shown."],
        ["expert_name", "text", "Name and surname", "Person filling or reviewing the row."],
        ["review_status", "text", REVIEW_STATUS_EN, "Minimal review tracking."],
        ["(notes)", "optional text", "Free observations", "Optional."],
        [""],
    ]
    return guide + [REPORT_COLUMNS_EN] + REPORT_EXAMPLES_EN


def eval_rows(lang: str) -> list[list[str]]:
    if lang == "es":
        guide = [
            ["campo", "formato", "opciones_o_ejemplo", "notas"],
            ["tipo_fila", "texto", "EJEMPLO|EXPERTOS", "Marca filas de ejemplo y la primera fila a completar."],
            ["nombre_pdf", "texto", "Mismo valor que en registro_informes", "Usar exactamente el mismo nombre."],
            ["nombre_fuente", "texto", "Mismo valor que en registro_informes", "Usar exactamente el mismo nombre."],
            ["codigo_unfc_final", "texto", "E1;F2;G3", "Código final del activo como conjunto."],
            ["categoria_e_final", "texto", "E1|E2|E3|E4|unknown", "Categoría final del eje E."],
            ["categoria_f_final", "texto", "F1|F2|F3|F4|unknown", "Categoría final del eje F."],
            ["categoria_g_final", "texto", "G1|G2|G3|G4|unknown", "Categoría final del eje G."],
            ["codigo_base_cantidad", "código numérico", BASE_CANTIDAD_GUIDE_ES, "Indica qué representan las cantidades."],
            ["cantidad_medida", "numero", "15700", "Cantidad medida en la base indicada por el código."],
            ["cantidad_indicada", "numero", "25500", "Cantidad indicada en la base indicada por el código."],
            ["cantidad_inferida", "numero", "16000", "Cantidad inferida en la base indicada por el código."],
            ["unidad", "texto", "t|kt|Mt|oz|kg", "Usar una única unidad por fila."],
            ["paginas_clave", "texto", "1-3,45-60", "Paginas que soportan la clasificacion."],
            ["justificacion_experto", "texto", "Explicacion breve", "Resumen de por qué la clasificacion final es la elegida."],
            [""],
        ]
        return guide + [EVAL_COLUMNS_ES] + EVAL_EXAMPLES_ES
    guide = [
        ["field", "format", "allowed_values_or_example", "notes"],
        ["row_type", "text", "EXAMPLE|EXPERTS", "Marks sample rows and the first row to complete."],
        ["pdf_name", "text", "Same value as in report_register", "Use exactly the same name."],
        ["source_name", "text", "Same value as in report_register", "Use exactly the same name."],
        ["unfc_code_final", "text", "E1;F2;G3", "Final code for the asset as a whole."],
        ["e_category_final", "text", "E1|E2|E3|E4|unknown", "Final E axis category."],
        ["f_category_final", "text", "F1|F2|F3|F4|unknown", "Final F axis category."],
        ["g_category_final", "text", "G1|G2|G3|G4|unknown", "Final G axis category."],
        ["quantity_basis_code", "numeric code", BASE_CANTIDAD_GUIDE_EN, "Explains what the quantity fields represent."],
        ["measured_quantity", "number", "15700", "Measured quantity in the basis indicated by the code."],
        ["indicated_quantity", "number", "25500", "Indicated quantity in the basis indicated by the code."],
        ["inferred_quantity", "number", "16000", "Inferred quantity in the basis indicated by the code."],
        ["unit", "text", "t|kt|Mt|oz|kg", "Use a single unit per row."],
        ["key_pages", "text", "1-3,45-60", "Pages supporting the classification."],
        ["expert_rationale", "text", "Short explanation", "Summary of why this final classification was chosen."],
        [""],
    ]
    return guide + [EVAL_COLUMNS_EN] + EVAL_EXAMPLES_EN


def evidence_rows(lang: str) -> list[list[str]]:
    guide = VARIABLE_GUIDE_ES if lang == "es" else VARIABLE_GUIDE_EN
    if lang == "es":
        top = [
            ["campo", "formato", "opciones_o_guia", "notas"],
            ["tipo_fila", "texto", "EJEMPLO|EXPERTOS", "Marca filas de ejemplo y la primera fila a completar."],
            ["nombre_pdf", "texto", "Mismo valor que en registro_informes", "Usar exactamente el mismo nombre."],
            ["nombre_fuente", "texto", "Mismo valor que en registro_informes", "Usar exactamente el mismo nombre."],
            ["eje", "texto", "E|F|G", "Eje al que contribuye la variable."],
            ["variable", "texto", "Usar lista sugerida o definir otra_variable", "Se pueden añadir todas las variables que el experto considere necesarias."],
            ["valor_numerico", "número", "Usar la escala indicada para cada variable", "El experto puede añadir tantas variables como necesite para justificar E, F y G."],
            ["paginas_evidencia", "texto", "12|13|14-16", "Las páginas son obligatorias; las notas y extractos son opcionales."],
            ["(extracto_o_nota)", "texto opcional", "fragmento breve o resumen", "Usar solo si ayuda a entender la justificación."],
            ["(comentario_experto)", "texto opcional", "explicación adicional", "Usar para razonar por qué una o varias variables empujan a E/F/G."],
            [""],
        ]
        return top + guide + [[""]] + [EVIDENCE_COLUMNS_ES] + build_evidence_examples("es")
    top = [
        ["field", "format", "allowed_values_or_guide", "notes"],
        ["row_type", "text", "EXAMPLE|EXPERTS", "Marks sample rows and the first row to complete."],
        ["pdf_name", "text", "Same value as in report_register", "Use exactly the same name."],
        ["source_name", "text", "Same value as in report_register", "Use exactly the same name."],
        ["axis", "text", "E|F|G", "Axis influenced by the variable."],
        ["variable", "text", "Use suggested list or define other_variable", "Experts may add any variables they need."],
        ["numeric_value", "number", "Use the numeric scale shown for each variable", "Experts may add as many variables as needed to justify E, F, and G."],
        ["evidence_pages", "text", "12|13|14-16", "Pages are mandatory; notes and excerpts are optional."],
        ["(excerpt_or_note)", "optional text", "short quote or summary", "Use only if it helps explain the justification."],
        ["(expert_comment)", "optional text", "additional explanation", "Use to explain why one or more variables drive E/F/G."],
        [""],
    ]
    return top + guide + [[""]] + [EVIDENCE_COLUMNS_EN] + build_evidence_examples("en")


def qa_rows(lang: str) -> list[list[str]]:
    if lang == "es":
        guide = [
            ["campo", "formato", "opciones_o_ejemplo", "notas"],
            ["tipo_fila", "texto", "EJEMPLO|EXPERTOS", "Marca filas de ejemplo y la primera fila a completar."],
            ["nombre_pdf", "texto", "Mismo valor que en registro_informes", "Usar exactamente el mismo nombre."],
            ["nombre_fuente", "texto", "Mismo valor que en registro_informes", "Usar exactamente el mismo nombre."],
            ["tipo_pregunta", "texto", QUESTION_TYPE_OPTIONS_ES, "Usar la categoría más concreta posible."],
            ["pregunta", "texto", "Siempre en inglés", "Asumimos preguntas en inglés para evaluar modelos."],
            ["respuesta_gold_corta", "texto", "Respuesta breve", "Debe ser la respuesta mínima correcta."],
            ["respuesta_gold_larga", "texto", "Respuesta desarrollada", "Puede incluir contexto adicional validado por el experto."],
            ["terminos_obligatorios", "texto", "term1|term2|term3", "Términos que deben aparecer en una respuesta correcta."],
            ["(alternativas_aceptables)", "texto opcional", "sinónimos o abreviaturas", "Opcional."],
            ["paginas_fuente", "texto", "12|13|14-16", "Páginas que justifican la respuesta gold."],
            ["(notas_revisor)", "texto opcional", "Observaciones libres", "Opcional."],
            [""],
        ]
        return guide + [QA_COLUMNS_ES] + QA_EXAMPLES_ES
    guide = [
        ["field", "format", "allowed_values_or_example", "notes"],
        ["row_type", "text", "EXAMPLE|EXPERTS", "Marks sample rows and the first row to complete."],
        ["pdf_name", "text", "Same value as in report_register", "Use exactly the same name."],
        ["source_name", "text", "Same value as in report_register", "Use exactly the same name."],
        ["question_type", "text", QUESTION_TYPE_OPTIONS_EN, "Use the most specific category possible."],
        ["question", "text", "Always in English", "We assume English questions for model evaluation."],
        ["gold_answer_short", "text", "Short answer", "Should be the minimal correct answer."],
        ["gold_answer_long", "text", "Expanded answer", "May include additional expert-validated context."],
        ["must_mention_terms", "text", "term1|term2|term3", "Terms that must appear in a correct answer."],
        ["(acceptable_alternatives)", "optional text", "synonyms or abbreviations", "Optional."],
        ["source_pages", "text", "12|13|14-16", "Pages supporting the gold answer."],
        ["(reviewer_notes)", "optional text", "Free observations", "Optional."],
        [""],
    ]
    return guide + [QA_COLUMNS_EN] + QA_EXAMPLES_EN


def sheet_map(lang: str) -> dict[str, list[list[str]]]:
    if lang == "es":
        return {
            "LEEME": readme_rows("es"),
            "registro_informes": report_rows("es"),
            "evaluacion_unfc": eval_rows("es"),
            "evidencia_ejes": evidence_rows("es"),
            "pares_qa": qa_rows("es"),
        }
    return {
        "README": readme_rows("en"),
        "report_register": report_rows("en"),
        "unfc_evaluation": eval_rows("en"),
        "axis_evidence": evidence_rows("en"),
        "qa_pairs": qa_rows("en"),
    }


def clean_dir(path: Path) -> None:
    if path.exists():
        for child in path.iterdir():
            if child.is_file():
                try:
                    child.unlink()
                except PermissionError:
                    # The workbook may be open in Excel; keep it and allow a _new file later.
                    continue
    path.mkdir(parents=True, exist_ok=True)


def build_set(lang: str, out_dir: Path, workbook_name: str) -> list[Path]:
    sheets = sheet_map(lang)
    generated: list[Path] = []
    for name, rows in sheets.items():
        csv_path = out_dir / f"{name}.csv"
        rows_to_csv(csv_path, rows)
        generated.append(csv_path)
    workbook_path = out_dir / workbook_name
    try:
        temp_path = out_dir / workbook_name.replace(".xlsx", ".__tmp__.xlsx")
        build_xlsx(temp_path, list(sheets.items()))
        os.replace(temp_path, workbook_path)
        generated.insert(0, workbook_path)
    except PermissionError:
        alt = out_dir / workbook_name.replace(".xlsx", "_new.xlsx")
        build_xlsx(alt, list(sheets.items()))
        generated.insert(0, alt)
    return generated


def build_all_templates() -> list[Path]:
    clean_dir(ES_DIR)
    clean_dir(EN_DIR)
    generated: list[Path] = []
    generated.extend(build_set("es", ES_DIR, "plantilla_expertos_unfc_es.xlsx"))
    generated.extend(build_set("en", EN_DIR, "unfc_expert_template_en.xlsx"))
    return generated


def main() -> None:
    generated = build_all_templates()
    print("Generated templates:")
    for path in generated:
        print(f" - {os.path.relpath(path, ROOT.parent)}")


if __name__ == "__main__":
    main()
